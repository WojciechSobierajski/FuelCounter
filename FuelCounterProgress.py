import openpyxl as xl
from datetime import datetime
import tkinter as tk
wb = xl.load_workbook('TemplateForFuelCounter.xlsx')
sheet = wb['Calculation']

#Defined functions
def displaying_calculated():
    dist=f"You've driven: {distance_driven} km"
    print(dist)
    trip_cost = f'It cost you: {total_cost} zł'
    print(trip_cost)
    cost_of_100km = f'100 km cost you: {cost_100km} zł'
    print(cost_of_100km)
    liters_per_100km= f'Fuel consumption is: {fuel_consumption} l/100 km'
    print(liters_per_100km)

score = ""
def displaying_validation_of_style():
    if fuel_consumption > 10:
        score = 'Too fast!'
    if fuel_consumption < 10 and fuel_consumption > 5:
        score = 'Fuel consumption is good!'
    if fuel_consumption < 5:
        score = 'Wow, very low !'
    print(score)


def displaying_Tkinter():
    def closing():
        root.destroy()
    root = tk.Tk()
    canvas = tk.Canvas(root, height=300, width=400, bg="green")
    canvas.pack()
    frame = tk.Frame(root, height=100, width=100, bg="white")
    frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)
    label = tk.Label(frame, text=(f"You've driven: {distance_driven} km"))
    label.place(relx=0.1, rely=0.1)
    label = tk.Label(frame, text=(f'It cost you: {total_cost} zl'))
    label.place(relx=0.1, rely=0.2)
    label = tk.Label(frame, text=(f'Fuel consumption is: {fuel_consumption} l/100km'))
    label.place(relx=0.1, rely=0.3)
    label = tk.Label(frame, text=(score))
    label.place(relx=0.1, rely=0.4)
    label = tk.Label(frame, text=(f'Overall car distance: {car_distance} km'))
    label.place(relx=0.1, rely=0.5)
    button = tk.Button(root, text='Click to continue', command=closing)
    button.pack()
    root.mainloop()


def filling_excel():
    cell = sheet.cell(sheet.max_row + 1, 1)
    cell.value = (datetime.date(datetime.now()))
    cell = sheet.cell(sheet.max_row, 2)
    cell.value = current_distance
    cell = sheet.cell(sheet.max_row, 3)
    cell.value = current_price
    cell = sheet.cell(sheet.max_row, 4)
    cell.value = current_liters
    cell = sheet.cell(sheet.max_row, 5)
    cell.value = distance_driven
    cell = sheet.cell(sheet.max_row, 6)
    cell.value = total_cost
    cell = sheet.cell(sheet.max_row, 7)
    cell.value = fuel_consumption
    cell = sheet.cell(sheet.max_row, 8)
    cell.value = cost_100km
    wb.save(f'{name}.xlsx')


#Beginning
question = ""
car_distance = 0
name = input("What's your name? ")
wb.save(f'{name}.xlsx')
wb = xl.load_workbook(f'{name}.xlsx')
sheet = wb['Calculation']
print(f'Hey {name}, welcome to my Fuel Counter! Have fun and calculate.')

while question != "yes":
# input_from_user
    current_distance = input('Current state of odometer (km) ')
    current_price = input('Price of 1 liter of gasoline ')
    current_price = current_price.replace(',','.')
    current_liters = input('How much liters did you fill? ')
    current_liters = current_liters.replace(',','.')
#calculation
    distance_driven = int(current_distance) - int(car_distance)
    total_cost = round(float(current_price) * float(current_liters), 2)
    fuel_consumption = round((float(current_liters) / float(distance_driven) * 100), 2)
    cost_100km = round(float(fuel_consumption * float(current_price)), 2)
    car_distance += int(distance_driven)
#Executing fuctions
    displaying_calculated()
    displaying_validation_of_style()
    displaying_Tkinter()
    filling_excel()


#loop question
    question = input("Are we finished? Yes or No? ").lower()
    if question == "yes":
        print("Thanks for using it! ")
        break
    if question == "no":
        print("Ok, let's go once again ")
    else:
        print("You were supposed to say Yes or No :P")
        question = input("Are we finished? Yes or No? ").lower()
